import json
import warnings
import logging
from pathlib import Path
from typing import Any, Union

from openpyxl import load_workbook

from api_integration.constants import (
    EXCEL_TO_RFQ_MAPPING,
    EXCEL_TO_RFQ_VALUES_MAPPING,
    RFQ_TO_DEFAULTS_MAPPING,
)
from api_integration.sevenrights.api.schemas.api_requests import RfqCreateRequest

logger = logging.getLogger(__name__)


def read_tender_excel(
    path: Union[str, Path], sheet_name: str = "TENDER"
) -> dict[str, str]:
    """
    Reads an Excel file, extracts data from the specified sheet,
    and returns a dictionary with keys from column A and values
    from column C (fallback to column B if C is empty/None).

    Args:
        path: Path to the Excel file (.xlsx, .xls, etc.)
        sheet_name: Name of the sheet to read (default: "TENDER")

    Returns:
        dict with str keys and str values, or {"error": "..."} if sheet is missing.
    """

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                action="ignore", message="Data Validation extension is not supported"
            )
            wb = load_workbook(path, data_only=True)
    except Exception as exc:
        return {"error": f"Failed to load workbook: {exc}"}

    if sheet_name not in wb.sheetnames:
        msg = f"Sheet '{sheet_name}' not found in {path}. Available sheets: {wb.sheetnames}"
        logger.info(msg)
        return {"error": msg}

    ws = wb[sheet_name]
    result: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0] if row and len(row) > 0 else None
        col_b = row[1] if row and len(row) > 1 else None
        col_c = row[2] if row and len(row) > 2 else None

        if key is None:
            continue

        key_str = str(key).strip()
        if not key_str:
            continue

        value = col_c if (col_c is not None and str(col_c).strip() != "") else col_b
        result[key_str] = str(value).strip() if value is not None else ""

    return result


def _parse_percent_range_last(value: str) -> float:
    """
    Extract the last numeric value from a percent range string and parse it to float.
    """
    parts = value.split("-")
    last = parts[-1].strip() if parts else value
    last = last.strip("%").strip()
    last = last.replace(",", ".")
    return float(last)


def _copy_price_rating_values(result: dict[str, Any]) -> None:
    """
    Ensure that all four traffic-light threshold fields are present in the result.
    If only one pair (price fields or rating fields) has values, copy them to the
    missing pair. Pre-existing values are never overwritten.

    Args:
        result: Dictionary with RFQ values, may contain some of the four fields.
    """
    price_green = result.get("price_green_finish_percent")
    price_yellow = result.get("price_yellow_finish_percent")
    rating_green = result.get("rating_green_finish_value")
    rating_yellow = result.get("rating_yellow_finish_value")

    has_price_values = price_green is not None or price_yellow is not None
    has_rating_values = rating_green is not None or rating_yellow is not None

    if has_price_values:
        # Price fields have values, ensure both exist and copy to rating fields
        if price_green is None:
            price_green = 0.0
        if price_yellow is None:
            price_yellow = 0.0

        result["price_green_finish_percent"] = price_green
        result["price_yellow_finish_percent"] = price_yellow

        if rating_green is None:
            result["rating_green_finish_value"] = price_green
        if rating_yellow is None:
            result["rating_yellow_finish_value"] = price_yellow
    elif has_rating_values:
        # Rating fields have values, ensure both exist and copy to price fields
        if rating_green is None:
            rating_green = 0.0
        if rating_yellow is None:
            rating_yellow = 0.0

        result["rating_green_finish_value"] = rating_green
        result["rating_yellow_finish_value"] = rating_yellow

        if price_green is None:
            result["price_green_finish_percent"] = rating_green
        if price_yellow is None:
            result["price_yellow_finish_percent"] = rating_yellow


def apply_excel_value_mappings(data: dict[str, str], sep: str = ";") -> dict[str, Any]:
    """
    Converts raw Excel string values to API-valid types using EXCEL_TO_RFQ_VALUES_MAPPING.

    For each key in data, if the key exists in EXCEL_TO_RFQ_VALUES_MAPPING and the value
    matches a known Excel value, it is replaced with the corresponding API value.
    If no exact match is found, the value is split by ';' and each part is mapped
    individually (unmapped parts are excluded). If at least one part is mapped:
    - all mapped values are lists -> flattened into a single list
    - otherwise -> single value if one mapped part, else list of mapped values

    Args:
        data: Dictionary with RFQ property names and raw Excel string values.
        sep: multi-valued Excel string values separator

    Returns:
        Dictionary with converted values.
    """
    result: dict[str, Any] = {}
    for key, value in data.items():
        if key not in EXCEL_TO_RFQ_VALUES_MAPPING:
            result[key] = value
            continue

        mapping = EXCEL_TO_RFQ_VALUES_MAPPING[key]
        if value in mapping:
            result[key] = mapping[value]
            continue

        parts = [part.strip() for part in value.split(sep)]
        mapped_parts: list[Any] = []
        for part in parts:
            if part in mapping:
                mapped_parts.append(mapping[part])

        if not mapped_parts:
            result[key] = value
        else:
            if all(isinstance(v, list) for v in mapped_parts):
                result[key] = [item for sublist in mapped_parts for item in sublist]
            else:
                result[key] = (
                    mapped_parts[0] if len(mapped_parts) == 1 else mapped_parts
                )

    # Type coercion for known numeric fields
    if "freight_spend_of_event" in result and isinstance(
        result["freight_spend_of_event"], str
    ):
        if result["freight_spend_of_event"].isdigit():
            result["freight_spend_of_event"] = int(result["freight_spend_of_event"])

    for field in (
        "price_green_finish_percent",
        "price_yellow_finish_percent",
        "rating_green_finish_value",
        "rating_yellow_finish_value",
    ):
        if field in result and isinstance(result[field], str):
            result[field] = _parse_percent_range_last(result[field])

    _copy_price_rating_values(result)

    return result


def remap_excel_to_rfq_properties(data: dict[str, str]) -> dict[str, str]:
    """
    Remaps dictionary keys from Excel field names to RFQ request property names
    using EXCEL_TO_RFQ_MAPPING from constants.py.

    Args:
        data: Dictionary returned by read_tender_excel()

    Returns:
        New dictionary with remapped keys where mapping exists.
    """
    result = {
        rfq_key: data[excel_key]
        for excel_key, rfq_key in EXCEL_TO_RFQ_MAPPING.items()
        if excel_key in data
    }
    return result


def merge_rfq_with_defaults(data: dict[str, str]) -> dict[str, str]:
    """
    Merges the remapped RFQ dictionary with RFQ_TO_DEFAULTS_MAPPING from constants.py.
    Missing fields are added from defaults; overlapping keys take precedence from defaults.

    Args:
        data: Dictionary returned by remap_excel_to_rfq_properties()

    Returns:
        Merged dictionary.
    """
    merged = dict(data)
    merged.update(RFQ_TO_DEFAULTS_MAPPING)
    return merged


def test_rfq_reader() -> None:
    """Run all RFQ reader tests and print statistics."""
    import sys

    default_path = (
        Path(__file__).resolve().parent.parent.parent
        / "samples"
        / "excel"
        / "Шаблон запроса создания тендера v1.4.xlsx"
    )
    file_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path

    if not Path(file_path).exists():
        print(f"File not found: {file_path}")
        sys.exit(1)

    stats: dict[str, Any] = {
        "rfq_fields_read": 0,
        "rfq_fields_mapped": 0,
        "rfq_fields_converted": 0,
        "rfq_template_valid": False,
        "errors": [],
    }

    # === Testing RFQ template functions ===
    print("=== Testing RFQ template functions ===")
    data = read_tender_excel(file_path)
    stats["rfq_fields_read"] = len(data)
    print(f"Read {stats['rfq_fields_read']} fields from RFQ template")

    rfq_data = remap_excel_to_rfq_properties(data)
    stats["rfq_fields_mapped"] = len(rfq_data)
    print(f"Mapped to {stats['rfq_fields_mapped']} RFQ properties")

    rfq_data = apply_excel_value_mappings(rfq_data)
    stats["rfq_fields_converted"] = len(rfq_data)
    print(f"Converted values for {stats['rfq_fields_converted']} properties")

    merged_data = merge_rfq_with_defaults(rfq_data)
    print(f"Merged with defaults, total fields: {len(merged_data)}")

    if "error" in merged_data:
        stats["errors"].append(merged_data["error"])
        print(f"RFQ template error: {merged_data['error']}")
    else:
        try:
            validated = RfqCreateRequest.model_validate(merged_data)
            stats["rfq_template_valid"] = True
            print("Pydantic validation: PASSED")
        except Exception as exc:
            stats["rfq_template_valid"] = False
            stats["errors"].append(str(exc))
            print(f"Pydantic validation: FAILED - {exc}")

    # === Statistics summary ===
    print("\n=== Test Statistics ===")
    print(f"RFQ fields read: {stats['rfq_fields_read']}")
    print(f"RFQ fields mapped: {stats['rfq_fields_mapped']}")
    print(f"RFQ fields converted: {stats['rfq_fields_converted']}")
    print(f"RFQ template valid: {stats['rfq_template_valid']}")
    print(f"Errors: {len(stats['errors'])}")
    if stats["errors"]:
        for err in stats["errors"]:
            print(f"  - {err}")


if __name__ == "__main__":
    test_rfq_reader()
