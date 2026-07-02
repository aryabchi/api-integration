from typing import Any, Union
from pathlib import Path

from openpyxl import load_workbook

from api_integration.constants import RFQ_DEFAULT_LOT_TEMPLATE_ID


def read_lot_excel(
    path: Union[str, Path], sheet_name: str = "Матрица ТЗ_РФ"
) -> dict[str, Any]:
    """
    Reads lot template Excel file.
    On success returns {"lot_template_id": RFQ_DEFAULT_LOT_TEMPLATE_ID}.
    On failure returns {"error": "..."}.
    """
    try:
        wb = load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            msg = f"Sheet '{sheet_name}' not found in {path}. Available sheets: {wb.sheetnames}"
            print(msg)
            return {"error": msg}
        # TODO: add data validation logic
        return {
            "lot_template_id": RFQ_DEFAULT_LOT_TEMPLATE_ID,
            "path": str(path),
        }
    except Exception as exc:
        return {"error": str(exc)}


def test_lot() -> None:
    """Run lot template tests."""
    lot_file_path = (
        Path(__file__).resolve().parent.parent.parent
        / "samples"
        / "excel"
        / "Копия ТЗ Самсунг Артем перезакуп Хабаровский край 19.06.2026.xlsx"
    )
    print("=== Testing lot template function ===")
    lot_data = read_lot_excel(lot_file_path)
    if "error" in lot_data:
        print(f"Lot template error: {lot_data['error']}")
    else:
        print(f"Lot template processed: {lot_data}")


if __name__ == "__main__":
    test_lot()
