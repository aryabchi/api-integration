import sys
import json
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any, List, Union

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise ImportError("openpyxl is required to read Excel files") from exc

src_path = str(Path(__file__).resolve().parent.parent)
if src_path not in sys.path:
    sys.path.append(src_path)

from constants import (
    ALLOWED_ATTACHMENT_FILE_EXTENSIONS,
    ALLOWED_ATTACHMENT_LOT_TEMPLATE_TERMS,
    ALLOWED_ATTACHMENT_RFQ_TEMPLATE_TERMS,
    EXCEL_TO_RFQ_MAPPING,
    EXCEL_TO_RFQ_VALUES_MAPPING,
    RFQ_TO_DEFAULTS_MAPPING,
)
from sevenrights.api.schemas.rfq import RfqCreateRequest
from excel.lot import read_lot_excel


@dataclass
class AttachmentTemplates:
    lot_template: List[str]
    rfq_template: List[str]


def find_attachment_templates(
    directory: Union[str, Path],
    extensions: tuple[str, ...] = ALLOWED_ATTACHMENT_FILE_EXTENSIONS,
) -> AttachmentTemplates:
    """
    Scans a directory for files with allowed extensions and classifies them
    as lot template or rfq template based on filename pattern matching.

    Args:
        directory: Path to the directory to search.
        extensions: Allowed file extensions (default from constants.py).

    Returns:
        AttachmentTemplates with lists of matching file names.
    """
    base = Path(directory)
    normalized_exts = tuple(ext.lower().lstrip(".") for ext in extensions)
    lot: List[str] = []
    rfq: List[str] = []

    if not base.is_dir():
        return AttachmentTemplates(lot_template=[], rfq_template=[])

    for path in sorted(base.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower().lstrip(".") not in normalized_exts:
            continue
        name = path.name.lower()
        if any(term in name for term in ALLOWED_ATTACHMENT_LOT_TEMPLATE_TERMS):
            lot.append(path.name)
        elif any(term in name for term in ALLOWED_ATTACHMENT_RFQ_TEMPLATE_TERMS):
            rfq.append(path.name)

    return AttachmentTemplates(lot_template=lot, rfq_template=rfq)


def validate_attachment_templates(templates: AttachmentTemplates) -> bool:
    """
    Validates AttachmentTemplates:
    - each list must contain exactly one file
    - file names must be different
    """
    if len(templates.lot_template) != 1 or len(templates.rfq_template) != 1:
        return False
    return templates.lot_template[0] != templates.rfq_template[0]


def process_attachments(
    directory: Union[str, Path],
    extensions: tuple[str, ...] = ALLOWED_ATTACHMENT_FILE_EXTENSIONS,
) -> dict[str, Any]:
    """
    End-to-end wrapper: find templates, validate, read lot and RFQ Excel, remap, merge.
    """
    result: dict[str, Any] = {"lot_template": None, "rfq_template": None, "error": None}

    try:
        templates = find_attachment_templates(directory, extensions)
        if not validate_attachment_templates(templates):
            exts = ", ".join(extensions)
            msg = (
                f"Attachments validation failed: must be two unique {exts} files. "
                f"Lot template filename should include one of: {', '.join(ALLOWED_ATTACHMENT_LOT_TEMPLATE_TERMS)}. "
                f"RFQ template filename should include one of: {', '.join(ALLOWED_ATTACHMENT_RFQ_TEMPLATE_TERMS)}."
            )
            print(msg)
            result["error"] = msg
            return result

        if not templates.rfq_template:
            msg = "No RFQ template file found"
            print(msg)
            result["error"] = msg
            return result

        if templates.lot_template:
            lot_path = Path(directory) / templates.lot_template[0]
            lot_result = read_lot_excel(lot_path)
            if "error" in lot_result:
                result["error"] = lot_result["error"]
                result["lot_template"] = None
            else:
                result["lot_template"] = lot_result

        rfq_path = Path(directory) / templates.rfq_template[0]
        data = read_tender_excel(rfq_path)
        if "error" in data:
            result["error"] = data["error"]
            return result

        rfq_data = remap_excel_to_rfq_properties(data)
        rfq_data = apply_excel_value_mappings(rfq_data)
        result["rfq_template"] = merge_rfq_with_defaults(rfq_data)

    except Exception as exc:
        result["error"] = str(exc)

    return result


def read_tender_excel(
    path: Union[str, Path], sheet_name: str = "TENDER"
) -> dict[str, str]:
    """
    Reads an Excel file, extracts data from the specified sheet,
    and returns a dictionary with keys from column A and values
    from column C (fallback to column B if C is empty/None).

    Args:
        path: Path to the Excel file (.xlsx, .xls, etc.)
        sheet_name: Name of the sheet to read (default: "TENDER")

    Returns:
        dict with str keys and str values, or {"error": "..."} if sheet is missing.
    """

    with warnings.catch_warnings():
        warnings.filterwarnings(
            action="ignore", message="Data Validation extension is not supported"
        )
        wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        msg = f"Sheet '{sheet_name}' not found in {path}. Available sheets: {wb.sheetnames}"
        print(msg)
        return {"error": msg}

    ws = wb[sheet_name]
    result: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0] if row and len(row) > 0 else None
        col_b = row[1] if row and len(row) > 1 else None
        col_c = row[2] if row and len(row) > 2 else None

        if key is None:
            continue

        key_str = str(key).strip()
        if not key_str:
            continue

        value = col_c if (col_c is not None and str(col_c).strip() != "") else col_b
        result[key_str] = str(value).strip() if value is not None else ""

    return result


def _parse_percent_range_last(value: str) -> float:
    parts = value.split("-")
    last = parts[-1].strip() if parts else value
    last = last.strip("%").strip()
    last = last.replace(",", ".")
    return float(last)


def apply_excel_value_mappings(data: dict[str, str]) -> dict[str, Any]:
    """
    Converts raw Excel string values to API-valid types using EXCEL_TO_RFQ_VALUES_MAPPING.

    For each key in data, if the key exists in EXCEL_TO_RFQ_VALUES_MAPPING and the value
    matches a known Excel value, it is replaced with the corresponding API value.
    If no exact match is found, the value is split by ';' and each part is mapped
    individually (unmapped parts are excluded). If at least one part is mapped:
    - all mapped values are lists -> flattened into a single list
    - otherwise -> single value if one mapped part, else list of mapped values
    If the input contains an 'error' key, returns a dict with only the error.

    Args:
        data: Dictionary with RFQ property names and raw Excel string values.

    Returns:
        Dictionary with converted values, or {'error': ...} if input contains an error.
    """
    if "error" in data:
        return {"error": data["error"]}

    result: dict[str, Any] = {}
    for key, value in data.items():
        if key not in EXCEL_TO_RFQ_VALUES_MAPPING:
            result[key] = value
            continue

        mapping = EXCEL_TO_RFQ_VALUES_MAPPING[key]
        if value in mapping:
            result[key] = mapping[value]
            continue

        parts = [part.strip() for part in value.split(";")]
        mapped_parts: list[Any] = []
        for part in parts:
            if part in mapping:
                mapped_parts.append(mapping[part])

        if not mapped_parts:
            result[key] = value
        else:
            if all(isinstance(v, list) for v in mapped_parts):
                result[key] = [item for sublist in mapped_parts for item in sublist]
            else:
                result[key] = (
                    mapped_parts[0] if len(mapped_parts) == 1 else mapped_parts
                )

    # Type coercion for known numeric fields
    if "freight_spend_of_event" in result and isinstance(
        result["freight_spend_of_event"], str
    ):
        if result["freight_spend_of_event"].isdigit():
            result["freight_spend_of_event"] = int(result["freight_spend_of_event"])

    for field in ("price_green_finish_percent", "price_yellow_finish_percent"):
        if field in result and isinstance(result[field], str):
            result[field] = _parse_percent_range_last(result[field])

    return result


def remap_excel_to_rfq_properties(data: dict[str, str]) -> dict[str, str]:
    """
    Remaps dictionary keys from Excel field names to RFQ request property names
    using EXCEL_TO_RFQ_MAPPING from constants.py.

    Args:
        data: Dictionary returned by read_tender_excel()

    Returns:
        New dictionary with remapped keys where mapping exists.
        If input contains 'error', it is retained in the output.
    """
    result = {
        rfq_key: data[excel_key]
        for excel_key, rfq_key in EXCEL_TO_RFQ_MAPPING.items()
        if excel_key in data
    }
    if "error" in data:
        result["error"] = data["error"]
    return result


def merge_rfq_with_defaults(data: dict[str, str]) -> dict[str, str]:
    """
    Merges the remapped RFQ dictionary with RFQ_TO_DEFAULTS_MAPPING from constants.py.
    Missing fields are added from defaults; overlapping keys take precedence from defaults.
    If the input contains an 'error' key, returns a dict with only the error.

    Args:
        data: Dictionary returned by remap_excel_to_rfq_properties()

    Returns:
        Merged dictionary, or {'error': ...} if input contains an error.
    """
    if "error" in data:
        return {"error": data["error"]}

    merged = dict(data)
    merged.update(RFQ_TO_DEFAULTS_MAPPING)
    return merged


def test_excel_reader() -> None:
    """Run all Excel reader tests and print statistics."""
    import sys

    default_path = (
        Path(__file__).resolve().parent.parent.parent
        / "samples"
        / "excel"
        / "Шаблон запроса создания тендера v1.3.xlsx"
        # / "Шаблон запроса создания тендера v1.2.xlsx"
        # / "Шаблон запроса создания тендера v1.3_remarks.xlsx"
    )
    file_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path

    if not Path(file_path).exists():
        print(f"File not found: {file_path}")
        sys.exit(1)

    stats: dict[str, Any] = {
        "rfq_fields_read": 0,
        "rfq_fields_mapped": 0,
        "rfq_fields_converted": 0,
        "rfq_template_valid": False,
        "lot_template_ok": False,
        "errors": [],
    }

    # === Testing RFQ template functions ===
    print("=== Testing RFQ template functions ===")
    data = read_tender_excel(file_path)
    stats["rfq_fields_read"] = len(data)
    print(f"Read {stats['rfq_fields_read']} fields from RFQ template")

    rfq_data = remap_excel_to_rfq_properties(data)
    stats["rfq_fields_mapped"] = len(rfq_data)
    print(f"Mapped to {stats['rfq_fields_mapped']} RFQ properties")

    rfq_data = apply_excel_value_mappings(rfq_data)
    stats["rfq_fields_converted"] = len(rfq_data)
    print(f"Converted values for {stats['rfq_fields_converted']} properties")

    merged_data = merge_rfq_with_defaults(rfq_data)
    print(f"Merged with defaults, total fields: {len(merged_data)}")

    if "error" in merged_data:
        stats["errors"].append(merged_data["error"])
        print(f"RFQ template error: {merged_data['error']}")
    else:
        try:
            validated = RfqCreateRequest.model_validate(merged_data)
            stats["rfq_template_valid"] = True
            print("Pydantic validation: PASSED")
        except Exception as exc:
            stats["rfq_template_valid"] = False
            stats["errors"].append(str(exc))
            print(f"Pydantic validation: FAILED - {exc}")

    # === Testing lot template function ===
    # print("\n=== Testing lot template function ===")
    # lot_file_path = (
    #     Path(__file__).resolve().parent.parent.parent
    #     / "samples"
    #     / "excel"
    #     / "Копия ТЗ Самсунг Артем перезакуп Хабаровский край 19.06.2026.xlsx"
    # )
    # lot_data = read_lot_excel(lot_file_path)
    # if "error" in lot_data:
    #     stats["errors"].append(lot_data["error"])
    #     print(f"Lot template error: {lot_data['error']}")
    # else:
    #     stats["lot_template_ok"] = True
    #     print(f"Lot template processed: {lot_data}")

    # === Testing end-to-end process_attachments pipeline ===
    print("\n=== Testing end-to-end process_attachments pipeline ===")
    samples_dir = Path(__file__).resolve().parent.parent.parent / "samples" / "excel"
    merged_data = process_attachments(samples_dir)
    print(f"lot_template: {'OK' if merged_data.get('lot_template') else 'None/Error'}")
    print(f"rfq_template: {'OK' if merged_data.get('rfq_template') else 'None/Error'}")
    print(f"error: {merged_data.get('error')}")

    with open(samples_dir / "rfq_excel.json", "w", encoding="utf-8") as f:
        json.dump(merged_data, f, ensure_ascii=False, indent=2)

    print("\n=== process_attachments result ===")
    print(json.dumps(merged_data, ensure_ascii=False, indent=2))

    # === Statistics summary ===
    print("\n=== Test Statistics ===")
    print(f"RFQ fields read: {stats['rfq_fields_read']}")
    print(f"RFQ fields mapped: {stats['rfq_fields_mapped']}")
    print(f"RFQ fields converted: {stats['rfq_fields_converted']}")
    print(f"RFQ template valid: {stats['rfq_template_valid']}")
    print(f"Lot template OK: {stats['lot_template_ok']}")
    print(f"Errors: {len(stats['errors'])}")
    if stats["errors"]:
        for err in stats["errors"]:
            print(f"  - {err}")


if __name__ == "__main__":
    test_excel_reader()
